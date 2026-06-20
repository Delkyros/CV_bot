import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_report(vagas_analisadas, output_path="vagas_filtradas.xlsx"):
    """
    Gera um relatório consolidado em Excel (ordenado por melhor match),
    estilizando a tabela final para uma visualização limpa e profissional.
    """
    print(f"\n[Reporter] Gerando relatório consolidado para {len(vagas_analisadas)} vagas analisadas...")
    
    if not vagas_analisadas:
        print("  [Aviso] Nenhuma vaga analisada com sucesso para gerar o relatório.")
        return False
        
    # Prepara a lista de registros formatados para o DataFrame
    registros = []
    for vaga in vagas_analisadas:
        # Formata pontos fortes e gaps como listas com marcadores (bullet points) para melhor leitura
        pontos_fortes_str = "\n".join([f"• {p}" for p in vaga.get("pontos_fortes", [])])
        gaps_str = "\n".join([f"• {g}" for g in vaga.get("gaps", [])])
        
        registros.append({
            "Nota de Match": vaga.get("nota_match", 0),
            "Título da Vaga": vaga.get("titulo_vaga", "N/A"),
            "Empresa": vaga.get("empresa", "N/A"),
            "Tipo de Contratação": vaga.get("tipo_contratacao", "N/A"),
            "Modelo de Trabalho": vaga.get("modelo_trabalho", "N/A"),
            "Localização": vaga.get("localizacao", "N/A"),
            "Link da Vaga": vaga.get("link_vaga", ""),
            "Pontos Fortes": pontos_fortes_str,
            "Gaps": gaps_str,
            "Veredicto": vaga.get("veredicto", "N/A")
        })
        
    # Cria o DataFrame e ordena por Nota de Match de forma decrescente
    df = pd.DataFrame(registros)
    df = df.sort_values(by="Nota de Match", ascending=False)
    
    # Define as colunas finais na ordem exata solicitada nas especificações
    colunas_ordenadas = [
        "Nota de Match", "Título da Vaga", "Empresa", "Tipo de Contratação",
        "Modelo de Trabalho", "Localização", "Link da Vaga",
        "Pontos Fortes", "Gaps", "Veredicto"
    ]
    df = df[colunas_ordenadas]
    
    try:
        # Salva o DataFrame no formato Excel usando Pandas
        df.to_excel(output_path, index=False, sheet_name="Vagas Filtradas")
        
        # Aplica estilizações avançadas utilizando o OpenPyXL para deixar o relatório profissional
        wb = load_workbook(output_path)
        ws = wb["Vagas Filtradas"]
        
        # Estilos: Cores, Fontes, Alinhamento, Bordas
        fonte_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        fonte_corpo = Font(name="Calibri", size=11)
        fonte_link = Font(name="Calibri", size=11, color="0563C1", underline="single")
        
        # Azul escuro corporativo para o cabeçalho
        preenchimento_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        
        # Cinza claro para linhas alternadas (efeito zebra) se necessário, ou bordas finas normais
        borda_fina = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        alinhamento_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
        alinhamento_esquerda = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        # Estilizando o Cabeçalho (Linha 1)
        ws.row_dimensions[1].height = 28
        for cell in ws[1]:
            cell.font = fonte_header
            cell.fill = preenchimento_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = borda_fina
            
        # Estilizando as Linhas de Dados
        for row in range(2, ws.max_row + 1):
            # Define altura confortável para acomodar células multi-linhas (gaps, pontos fortes)
            ws.row_dimensions[row].height = 65
            
            for col in range(1, len(colunas_ordenadas) + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = fonte_corpo
                cell.border = borda_fina
                
                # Alinhamento e formatação específica por tipo de coluna
                if col == 1: # Nota de Match
                    cell.alignment = alinhamento_centro
                    # Destaca notas altas com cores suaves de fundo
                    nota = cell.value
                    if isinstance(nota, (int, float)):
                        if nota >= 80: # Ótimo Match
                            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Verde claro
                            cell.font = Font(name="Calibri", size=11, bold=True, color="375623")
                        elif nota >= 50: # Match Médio
                            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Amarelo claro
                            cell.font = Font(name="Calibri", size=11, bold=True, color="7F6000")
                        else: # Match Baixo
                            cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Vermelho claro
                            cell.font = Font(name="Calibri", size=11, bold=True, color="C65911")
                elif col == 7: # Link da Vaga (Formatado como link clicável)
                    cell.alignment = alinhamento_esquerda
                    cell.font = fonte_link
                    if cell.value:
                        cell.hyperlink = cell.value
                else:
                    cell.alignment = alinhamento_esquerda
                    
        # Redimensionamento automático da largura das colunas
        larguras_padrao = {
            1: 15, # Nota de Match
            2: 30, # Título da Vaga
            3: 20, # Empresa
            4: 20, # Tipo de Contratação
            5: 20, # Modelo de Trabalho
            6: 28, # Localização
            7: 25, # Link da Vaga
            8: 35, # Pontos Fortes
            9: 35, # Gaps
            10: 40  # Veredicto
        }
        
        for col_idx, largura in larguras_padrao.items():
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = largura
            
        # Salva as alterações de estilo do openpyxl
        wb.save(output_path)
        print(f"[Reporter] Sucesso! Relatório salvo e estilizado em '{output_path}'.")
        return True
        
    except Exception as e:
        print(f"  [Erro Reporter] Falha ao gerar relatório em Excel: {e}")
        return False
